from __future__ import annotations

import io
import re
from pathlib import Path
from typing import Any
from unicodedata import east_asian_width
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


EXPORT_COLUMNS = (
    ("用例编号", "case_id"),
    ("用例标题", "title"),
    ("测试模块", "module"),
    ("前置条件", "preconditions"),
    ("执行步骤", "steps"),
    ("测试输入", "test_input"),
    ("预期结果", "expected_results"),
    ("用例级别", "priority"),
)

_COLUMN_WIDTHS = (16, 34, 22, 34, 46, 34, 46, 12)
_ILLEGAL_XML_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")
_STEP_NUMBER_RE = re.compile(r"^\s*(?:\d+\s*[.、)]|[（(]\s*\d+\s*[）)])\s*")
_UNSAFE_FORMULA_RE = re.compile(r"^[\s\t\r\n]*[=+\-@]")
_UNSAFE_FILENAME_RE = re.compile(r'[\x00-\x1F<>:"/\\|?*]+')


def _excel_text(value: Any) -> str:
    if value is None:
        return ""
    text = _ILLEGAL_XML_RE.sub("", str(value))
    return f"'{text}" if _UNSAFE_FORMULA_RE.match(text) else text


def _text_lines(value: Any) -> list[str]:
    values = value if isinstance(value, list) else [value]
    return [
        text
        for item in values
        if (text := _excel_text(item).strip())
    ]


def _numbered_lines(values: list[Any]) -> str:
    lines = _text_lines(values)
    return "\n".join(
        line if _STEP_NUMBER_RE.match(line) else f"{index}. {line}"
        for index, line in enumerate(lines, start=1)
    )


def _estimated_wrapped_lines(value: Any, column_width: int) -> int:
    text = str(value or "")
    available_width = max(1, int(column_width) - 2)
    line_count = 0
    for line in text.split("\n") or [""]:
        display_width = sum(
            2 if east_asian_width(character) in {"W", "F"} else 1
            for character in line
        )
        line_count += max(1, (display_width + available_width - 1) // available_width)
    return line_count


def _project_case(test_case: dict[str, Any]) -> list[str]:
    actions: list[Any] = []
    expected_results: list[Any] = []
    raw_steps = test_case.get("steps")
    for step in raw_steps if isinstance(raw_steps, list) else []:
        if isinstance(step, dict):
            actions.append(step.get("action"))
            expected_results.append(step.get("expected"))
        else:
            actions.append(step)

    values = {
        "case_id": test_case.get("case_id"),
        "title": test_case.get("title"),
        "module": test_case.get("module"),
        "preconditions": "\n".join(_text_lines(test_case.get("preconditions"))),
        "steps": _numbered_lines(actions),
        "test_input": test_case.get("test_input"),
        "expected_results": _numbered_lines(expected_results),
        "priority": test_case.get("priority"),
    }
    return [_excel_text(values[field]) for _header, field in EXPORT_COLUMNS]


def build_test_cases_excel(test_cases: list[dict[str, Any]]) -> bytes:
    if not test_cases:
        raise ValueError("没有可导出的测试用例")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "测试用例"
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    worksheet.append([header for header, _field in EXPORT_COLUMNS])

    for test_case in test_cases:
        worksheet.append(_project_case(test_case))

    header_fill = PatternFill(fill_type="solid", fgColor="8EDB82")
    header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="111827")
    body_font = Font(name="Microsoft YaHei", size=10, color="1F2937")
    separator = Side(style="thin", color="D9E2F3")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="medium", color="4F8A47"))
    worksheet.row_dimensions[1].height = 28

    for column_index, width in enumerate(_COLUMN_WIDTHS, start=1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    priority_fills = {
        "P0": PatternFill(fill_type="solid", fgColor="FCE8E6"),
        "P1": PatternFill(fill_type="solid", fgColor="FFF4CE"),
        "P2": PatternFill(fill_type="solid", fgColor="E2F0D9"),
    }
    for row_index in range(2, worksheet.max_row + 1):
        row_cells = worksheet[row_index]
        max_lines = 1
        for cell, column_width in zip(row_cells, _COLUMN_WIDTHS):
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=separator)
            max_lines = max(
                max_lines,
                _estimated_wrapped_lines(cell.value, column_width),
            )
        priority_cell = row_cells[-1]
        priority_cell.alignment = Alignment(horizontal="center", vertical="top")
        priority_cell.fill = priority_fills.get(str(priority_cell.value or "").upper(), PatternFill())
        worksheet.row_dimensions[row_index].height = min(360, max(24, 18 * max_lines))

    last_column = get_column_letter(len(EXPORT_COLUMNS))
    worksheet.auto_filter.ref = f"A1:{last_column}{worksheet.max_row}"
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.print_title_rows = "1:1"

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_cases_export_filename(source_filename: str, run_id: int) -> tuple[str, str]:
    source_stem = Path(str(source_filename or "")).stem.strip()
    safe_stem = _UNSAFE_FILENAME_RE.sub("_", source_stem).strip(" ._")
    display_name = f"{safe_stem or f'Run_{run_id}'}_测试用例.xlsx"
    ascii_name = f"test_cases_run_{run_id}.xlsx"
    disposition = f"attachment; filename={ascii_name}; filename*=UTF-8''{quote(display_name)}"
    return display_name, disposition
