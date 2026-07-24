from __future__ import annotations

import io
import json
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..postprocess.case_access import (
    case_id as case_access_id,
    case_priority,
    case_step_lines,
    case_text_field,
    case_text_list_field,
    case_text_list_value,
)
from ..postprocess.streaming_execution_plan_ordering import execution_group_order_rank


_GROUP_NAMES = {
    "main_smoke": "主链路冒烟",
    "permission": "权限与鉴权",
    "exception": "异常与失败路径",
    "boundary": "边界数据",
    "independent_functional": "独立功能",
    "independent": "独立功能",
    "display": "展示与查询",
    "unknown": "手工执行顺序",
}
_ILLEGAL_XML_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")


def _text(value: Any) -> str:
    return str(value or "").strip()


def _safe_int(value: Any, default: int = 0) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return int(default)


def _excel_text(value: Any) -> str:
    if isinstance(value, list):
        value = "\n".join(_text(item) for item in value if _text(item))
    return _ILLEGAL_XML_RE.sub("", _text(value))


def _case_id(case: dict[str, Any], index: int) -> str:
    return case_access_id(case) or f"TC-{index:03d}"


def _execution_group(case: dict[str, Any]) -> str:
    group = _text(case.get("execution_group")).lower()
    return group or "unknown"


def _chain_id(case: dict[str, Any], group: str) -> str:
    chain_id = _text(case.get("chain_id"))
    if chain_id:
        return chain_id
    if group == "main_smoke":
        return "main_smoke_chain"
    return f"{group}_suite"


def _suite_name(group: str, chain_id: str) -> str:
    if chain_id == "main_smoke_chain":
        return _GROUP_NAMES["main_smoke"]
    return _GROUP_NAMES.get(group, chain_id or _GROUP_NAMES["unknown"])


def _case_sort_key(case: dict[str, Any], fallback_index: int) -> tuple[int, str]:
    sequence = _safe_int(case.get("execution_sequence"), 0)
    if sequence <= 0:
        sequence = fallback_index
    return sequence, _case_id(case, fallback_index)


def _case_action(case: dict[str, Any], steps: list[str]) -> str:
    step_action = _text(steps[0]) if steps else ""
    if step_action:
        return step_action
    return _text(case.get("action"))


def _case_transition_action(case: dict[str, Any]) -> str:
    return _text(case.get("transition_action")) or _text(case.get("action"))


def _run_mode(group: str) -> str:
    if group == "main_smoke":
        return "sequential"
    if group == "unknown":
        return "manual_sequential"
    return "isolated"


def _has_value(value: Any) -> bool:
    return value not in (None, "", [])


def parse_generated_cases_payload(payload: Any) -> list[dict[str, Any]]:
    """Parse persisted generated_result shapes without changing route compatibility."""
    data = payload
    if isinstance(payload, str) and payload.strip():
        try:
            data = json.loads(payload)
        except Exception:
            return []
    if isinstance(data, dict):
        for key in ("cases", "generated_result", "final_cases"):
            value = data.get(key)
            if isinstance(value, list):
                data = value
                break
            if isinstance(value, str) and value.strip():
                parsed = parse_generated_cases_payload(value)
                if parsed:
                    return parsed
        else:
            return []
    if not isinstance(data, list):
        return []
    return [dict(item) for item in data if isinstance(item, dict)]


def _build_case_ref(case: dict[str, Any], *, fallback_index: int, suite_order: int) -> dict[str, Any]:
    case_id = _case_id(case, fallback_index)
    depends_on = case_text_list_value(case.get("depends_on"), split_lines=True)
    steps = case_step_lines(case)
    expected_result = case_text_field(case, "expected_result")
    return {
        "case_id": case_id,
        "suite_order": int(suite_order),
        "execution_sequence": _safe_int(case.get("execution_sequence"), suite_order),
        "description": case_text_field(case, "description"),
        "test_module": case_text_field(case, "test_module"),
        "priority": case_priority(case, prefer_final=True),
        "role": _text(case.get("role")),
        "session_key": _text(case.get("session_key")),
        "depends_on": depends_on,
        "fixture_key": _text(case.get("fixture_key")),
        "setup_hint": _text(case.get("setup_hint")),
        "teardown_hint": _text(case.get("teardown_hint")),
        "preconditions": case_text_list_field(case, "preconditions", split_lines=True),
        "steps": steps,
        "test_input": case_text_field(case, "test_input"),
        "expected_result": expected_result,
        "source_state": _text(case.get("source_state")),
        "target_state": _text(case.get("target_state")),
        "action": _case_action(case, steps),
        "transition_action": _case_transition_action(case),
        "runnable": bool(case_id and steps and expected_result),
    }


def build_execution_suite(
    cases_payload: Any,
    *,
    workflow_absence_declared: bool = False,
) -> dict[str, Any]:
    """Build a run-friendly suite view from already generated case metadata."""
    cases = parse_generated_cases_payload(cases_payload)
    grouped: dict[str, dict[str, Any]] = {}
    case_ids = {_case_id(case, index) for index, case in enumerate(cases, start=1)}
    metadata_fields = ("execution_group", "execution_sequence", "role", "session_key")
    metadata_counts = {
        field: int(sum(1 for case in cases if _has_value(case.get(field))))
        for field in metadata_fields
    }
    has_any_execution_metadata = any(metadata_counts.values())
    complete_execution_metadata = bool(cases) and all(count == len(cases) for count in metadata_counts.values())

    for index, case in enumerate(cases, start=1):
        group = _execution_group(case)
        chain_id = _chain_id(case, group)
        suite = grouped.setdefault(
            chain_id,
            {
                "suite_id": chain_id,
                "suite_name": _suite_name(group, chain_id),
                "execution_group": group,
                "run_mode": _run_mode(group),
                "group_setup": _text(case.get("group_setup")),
                "group_teardown": _text(case.get("group_teardown")),
                "cases": [],
            },
        )
        if not suite.get("group_setup") and _text(case.get("group_setup")):
            suite["group_setup"] = _text(case.get("group_setup"))
        if not suite.get("group_teardown") and _text(case.get("group_teardown")):
            suite["group_teardown"] = _text(case.get("group_teardown"))
        suite["cases"].append((index, case))

    suites: list[dict[str, Any]] = []
    for suite in grouped.values():
        raw_cases = sorted(
            list(suite.pop("cases") or []),
            key=lambda item: _case_sort_key(item[1], item[0]),
        )
        case_refs = [
            _build_case_ref(case, fallback_index=fallback_index, suite_order=order)
            for order, (fallback_index, case) in enumerate(raw_cases, start=1)
        ]
        dependency_values = [dep for case_ref in case_refs for dep in case_ref.get("depends_on", [])]
        missing_dependencies = sorted({dep for dep in dependency_values if dep and dep not in case_ids})
        roles = sorted({case_ref["role"] for case_ref in case_refs if case_ref.get("role")})
        fixture_keys = sorted({case_ref["fixture_key"] for case_ref in case_refs if case_ref.get("fixture_key")})
        suite["case_count"] = int(len(case_refs))
        suite["roles"] = roles
        suite["fixture_keys"] = fixture_keys
        suite["missing_dependencies"] = missing_dependencies
        suite["runnable"] = bool(case_refs) and not missing_dependencies and all(
            bool(case_ref.get("runnable")) for case_ref in case_refs
        )
        suite_warnings: list[str] = []
        if suite.get("execution_group") == "unknown":
            suite_warnings.append("缺少执行分组元数据，已按原始生成顺序组织为手工执行套件")
        if missing_dependencies:
            suite_warnings.append("存在缺失依赖用例，需先补齐依赖后再自动执行")
        if suite.get("execution_group") == "main_smoke" and len(case_refs) < 2:
            suite_warnings.append("主链路用例不足 2 条，无法形成端到端闭环")
        suite["warnings"] = suite_warnings
        suite["cases"] = case_refs
        suites.append(suite)

    suites.sort(
        key=lambda suite: (
            execution_group_order_rank(_text(suite.get("execution_group")).lower()),
            _safe_int((suite.get("cases") or [{}])[0].get("execution_sequence"), 0),
            _text(suite.get("suite_id")),
        )
    )

    flat_run_order: list[dict[str, Any]] = []
    for suite_index, suite in enumerate(suites, start=1):
        for case_ref in suite.get("cases") or []:
            flat_run_order.append(
                {
                    "run_index": int(len(flat_run_order) + 1),
                    "suite_index": int(suite_index),
                    "suite_id": suite.get("suite_id"),
                    "suite_name": suite.get("suite_name"),
                    "case_id": case_ref.get("case_id"),
                    "execution_sequence": case_ref.get("execution_sequence"),
                    "depends_on": list(case_ref.get("depends_on") or []),
                    "role": case_ref.get("role"),
                    "session_key": case_ref.get("session_key"),
                    "fixture_key": case_ref.get("fixture_key"),
                }
            )

    main_suite = next((suite for suite in suites if suite.get("execution_group") == "main_smoke"), None)
    missing_dependency_count = int(
        sum(len(suite.get("missing_dependencies") or []) for suite in suites)
    )
    linear_executable = bool(main_suite and main_suite.get("runnable") and int(main_suite.get("case_count") or 0) >= 2)
    warnings: list[str] = []
    if not cases:
        warnings.append("没有解析到可导出的测试用例")
    elif not has_any_execution_metadata:
        warnings.append("历史结果缺少执行元数据，仅能按原始用例顺序手工执行")
    else:
        missing_fields = [field for field, count in metadata_counts.items() if count < len(cases)]
        if missing_fields:
            warnings.append("部分用例缺少执行元数据：" + ", ".join(missing_fields))
    if cases and not main_suite and not workflow_absence_declared:
        warnings.append("缺少 main_smoke 主链，无法确认端到端线性执行顺序")
    if missing_dependency_count:
        warnings.append(f"存在 {missing_dependency_count} 个缺失依赖引用")

    if workflow_absence_declared and cases and complete_execution_metadata and missing_dependency_count == 0:
        readiness = "independent_ready"
    elif linear_executable and complete_execution_metadata and missing_dependency_count == 0:
        readiness = "ready"
    elif cases and not has_any_execution_metadata:
        readiness = "legacy_manual"
    elif cases:
        readiness = "partial"
    else:
        readiness = "empty"

    return {
        "kind": "execution_suite",
        "version": "execution-suite-v1",
        "case_count": int(len(cases)),
        "suite_count": int(len(suites)),
        "runnable_suite_count": int(sum(1 for suite in suites if bool(suite.get("runnable")))),
        "linear_executable": linear_executable,
        "workflow_absence_declared": bool(workflow_absence_declared),
        "execution_readiness": readiness,
        "main_suite_id": _text((main_suite or {}).get("suite_id")),
        "metadata_quality": {
            "complete_execution_metadata": complete_execution_metadata,
            "has_any_execution_metadata": bool(has_any_execution_metadata),
            "field_counts": metadata_counts,
            "missing_dependency_count": missing_dependency_count,
        },
        "warnings": warnings,
        "suites": suites,
        "flat_run_order": flat_run_order,
    }


def _append_rows(ws, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([_excel_text(value) for value in row])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for column_index, header in enumerate(headers, start=1):
        letter = get_column_letter(column_index)
        ws.column_dimensions[letter].width = max(12, min(48, len(header) * 2 + 8))
        for cell in ws[letter]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def append_execution_suite_worksheets(
    wb: Workbook,
    suite_payload: Any,
    *,
    use_active_sheet: bool = False,
) -> None:
    suite = suite_payload if isinstance(suite_payload, dict) else build_execution_suite(suite_payload)
    summary_ws = wb.active if use_active_sheet else wb.create_sheet("执行套件")
    summary_ws.title = "执行套件"
    run_ws = wb.create_sheet("执行顺序")

    summary_rows: list[list[Any]] = []
    for item in suite.get("suites") or []:
        summary_rows.append(
            [
                item.get("suite_id"),
                item.get("suite_name"),
                item.get("execution_group"),
                item.get("run_mode"),
                item.get("case_count"),
                ", ".join(item.get("roles") or []),
                ", ".join(item.get("fixture_keys") or []),
                item.get("group_setup"),
                item.get("group_teardown"),
                "是" if item.get("runnable") else "否",
                ", ".join(item.get("missing_dependencies") or []),
                ", ".join(item.get("warnings") or []),
            ]
        )
    _append_rows(
        summary_ws,
        ["套件ID", "套件名称", "执行分组", "执行模式", "用例数", "角色", "Fixture", "组级准备", "组级清理", "可直接执行", "缺失依赖", "诊断提示"],
        summary_rows,
    )

    run_rows: list[list[Any]] = []
    for suite_item in suite.get("suites") or []:
        for case in suite_item.get("cases") or []:
            run_rows.append(
                [
                    len(run_rows) + 1,
                    suite_item.get("suite_id"),
                    suite_item.get("suite_name"),
                    case.get("execution_sequence"),
                    case.get("case_id"),
                    case.get("description"),
                    case.get("test_module"),
                    case.get("priority"),
                    case.get("role"),
                    case.get("session_key"),
                    case.get("depends_on"),
                    case.get("setup_hint"),
                    case.get("steps"),
                    case.get("expected_result"),
                    case.get("teardown_hint"),
                ]
            )
    _append_rows(
        run_ws,
        ["全局顺序", "套件ID", "套件名称", "组内顺序", "用例ID", "用例标题", "模块", "优先级", "角色", "会话", "依赖用例", "准备说明", "测试步骤", "预期结果", "清理说明"],
        run_rows,
    )


def convert_execution_suite_to_excel(suite_payload: Any) -> bytes:
    wb = Workbook()
    append_execution_suite_worksheets(wb, suite_payload, use_active_sheet=True)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


__all__ = [
    "build_execution_suite",
    "append_execution_suite_worksheets",
    "convert_execution_suite_to_excel",
    "parse_generated_cases_payload",
]
