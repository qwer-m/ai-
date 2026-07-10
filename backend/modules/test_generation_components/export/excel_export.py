"""
测试用例导出组件。
该组件负责把测试用例 JSON 转换为 Excel/CSV 二进制内容。
"""

from __future__ import annotations

import ast
import io
import re
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..execution.execution_suite import append_execution_suite_worksheets, build_execution_suite
from ..postprocess.case_access import case_priority, case_text_list_value, case_text_value, case_value

# 中文注释：内部契约列仍保留给调试/回归使用，默认导出不直接暴露这些字段。
INTERNAL_EXPORT_COLUMNS = [
    "id",
    "description",
    "test_module",
    "preconditions",
    "steps",
    "test_input",
    "expected_result",
    "priority",
    "priority_final",
    "workflow_id",
    "source_state",
    "action",
    "target_state",
    "path_type",
    "blocking",
    "destructive",
    "can_advance_main_flow",
    "execution_group",
    "execution_sequence",
    "role",
    "session_key",
    "depends_on",
    "fixture_key",
    "setup_hint",
    "teardown_hint",
]
EXPORT_COLUMNS = INTERNAL_EXPORT_COLUMNS

CANONICAL_CASE_EXPORT_FIELDS = (
    "id",
    "description",
    "test_module",
    "preconditions",
    "steps",
    "test_input",
    "expected_result",
    "priority",
    "priority_final",
)

PUBLIC_EXPORT_COLUMNS = [
    ("用例ID", "id"),
    ("用例标题", "description"),
    ("所属模块", "test_module"),
    ("前置条件", "preconditions"),
    ("测试步骤", "steps"),
    ("测试数据", "test_input"),
    ("预期结果", "expected_result"),
    ("优先级", "priority"),
]

PUBLIC_COLUMN_WIDTHS = {
    "用例ID": 14,
    "用例标题": 36,
    "所属模块": 22,
    "前置条件": 34,
    "测试步骤": 42,
    "测试数据": 28,
    "预期结果": 44,
    "优先级": 12,
}

# 中文注释：Excel XML 不允许的控制字符，需在写入前清洗。
_ILLEGAL_XML_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")
_STEP_NUMBER_RE = re.compile(r"^\s*(?:\d+\s*[\.\)、)]|[（(]\s*\d+\s*[）)])\s*")


def _sanitize_excel_text(value: Any) -> str:
    """清洗单元格文本，避免 openpyxl 因非法字符报错。"""
    if value is None:
        return ""
    text = str(value)
    return _ILLEGAL_XML_RE.sub("", text)


def _has_export_value(value: Any) -> bool:
    return value not in (None, "", [])


def _fill_case_alias_fields(row: dict[str, Any]) -> None:
    for field in CANONICAL_CASE_EXPORT_FIELDS:
        if _has_export_value(row.get(field)):
            continue
        value = case_value(row, field, None)
        if _has_export_value(value):
            row[field] = value


def _format_steps_for_export(values: list[Any]) -> str:
    step_list = case_text_list_value(values)
    return "\n".join(
        step if _STEP_NUMBER_RE.match(step) else f"{i}. {step}"
        for i, step in enumerate(step_list, 1)
    )


def _presentation_order(row: dict[str, Any], fallback: int) -> int:
    for key in ("presentation_order", "presentationOrder", "display_order", "displayOrder"):
        try:
            value = int(row.get(key) or 0)
        except (TypeError, ValueError):
            value = 0
        if value > 0:
            return int(value)
    return int(fallback)


def _sort_rows_for_public_presentation(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not any(_presentation_order(row, 0) > 0 for row in rows if isinstance(row, dict)):
        return rows
    return [
        row
        for _order, _index, row in sorted(
            (
                (_presentation_order(row, 1_000_000 + index), index, row)
                for index, row in enumerate(rows, start=1)
            ),
            key=lambda item: (item[0], item[1]),
        )
    ]


def _normalize_rows(json_data: list | dict) -> list[dict[str, Any]]:
    """把输入数据统一成可导出的字典列表，并处理 steps/preconditions 格式。"""
    data: Any = json_data
    if isinstance(json_data, dict):
        if "error" in json_data:
            data = [{"error": json_data["error"]}]
        else:
            data = [json_data]
    if not isinstance(data, list):
        data = [data]

    rows: list[dict[str, Any]] = []
    for item in data:
        if not isinstance(item, dict):
            rows.append({"raw": str(item)})
            continue

        row = item.copy()
        _fill_case_alias_fields(row)

        pre = row.get("preconditions")
        if isinstance(pre, list):
            row["preconditions"] = "\n".join(case_text_list_value(pre))
        elif isinstance(pre, str) and pre.strip().startswith("[") and pre.strip().endswith("]"):
            try:
                val = ast.literal_eval(pre)
                if isinstance(val, list):
                    row["preconditions"] = "\n".join(case_text_list_value(val))
            except Exception:
                pass

        steps = row.get("steps")
        if isinstance(steps, list):
            row["steps"] = _format_steps_for_export(steps)
        elif isinstance(steps, str) and steps.strip().startswith("[") and steps.strip().endswith("]"):
            try:
                val = ast.literal_eval(steps)
                if isinstance(val, list):
                    row["steps"] = _format_steps_for_export(val)
            except Exception:
                pass

        depends_on = row.get("depends_on")
        if isinstance(depends_on, list):
            row["depends_on"] = "\n".join(case_text_list_value(depends_on))

        rows.append(row)
    return _sort_rows_for_public_presentation(rows)


def _project_public_rows(rows: list[dict[str, Any]]) -> tuple[list[str], list[dict[str, Any]]]:
    if rows and all("error" in row and not any(row.get(field) for _header, field in PUBLIC_EXPORT_COLUMNS) for row in rows):
        return ["错误信息"], [{"错误信息": row.get("error", "")} for row in rows]
    if rows and all("raw" in row and not any(row.get(field) for _header, field in PUBLIC_EXPORT_COLUMNS) for row in rows):
        return ["原始内容"], [{"原始内容": row.get("raw", "")} for row in rows]

    headers = [header for header, _field in PUBLIC_EXPORT_COLUMNS]
    projected: list[dict[str, Any]] = []
    for row in rows:
        item: dict[str, Any] = {}
        for header, field in PUBLIC_EXPORT_COLUMNS:
            value = row.get(field, "")
            if field == "priority":
                value = case_priority(row, prefer_final=True)
            item[header] = value
        projected.append(item)
    return headers, projected


def _project_internal_rows(rows: list[dict[str, Any]]) -> tuple[list[str], list[dict[str, Any]]]:
    return list(INTERNAL_EXPORT_COLUMNS), rows


def _apply_worksheet_layout(ws, headers: list[str]) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for column_index, header in enumerate(headers, start=1):
        letter = get_column_letter(column_index)
        ws.column_dimensions[letter].width = PUBLIC_COLUMN_WIDTHS.get(header, 20)
        for cell in ws[letter]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _build_excel_bytes(rows: list[dict[str, Any]], *, include_internal_fields: bool = False) -> bytes:
    """使用 openpyxl 按稳定列写 Excel，避免退化成单列 CSV。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例" if not include_internal_fields else "Test Cases"

    headers, projected_rows = (
        _project_internal_rows(rows)
        if include_internal_fields
        else _project_public_rows(rows)
    )

    ws.append(headers)
    for row in projected_rows:
        excel_row = [_sanitize_excel_text(row.get(col, "")) for col in headers]
        ws.append(excel_row)
    _apply_worksheet_layout(ws, headers)
    _append_execution_sheets_if_available(wb, rows)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def _looks_like_generated_case(row: dict[str, Any]) -> bool:
    return any(
        row.get(field) not in (None, "", [])
        for field in (
            "id",
            "case_id",
            "description",
            "steps",
            "expected_result",
            "execution_group",
            "execution_sequence",
            "depends_on",
        )
    )


def _append_execution_sheets_if_available(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    if not any(_looks_like_generated_case(row) for row in rows):
        return
    suite = build_execution_suite(rows)
    if int(suite.get("case_count") or 0) <= 0:
        return
    append_execution_suite_worksheets(wb, suite)


def convert_json_to_excel(json_data: list | dict, *, include_internal_fields: bool = False) -> bytes:
    """
    将测试用例 JSON 转换为 Excel 字节流。
    优先返回 xlsx；仅在极端异常时回退 CSV。
    """
    rows = _normalize_rows(json_data)

    try:
        return _build_excel_bytes(rows, include_internal_fields=include_internal_fields)
    except Exception as e:
        # 中文注释：保留兜底，但显式打印异常，便于排查为何回退 CSV。
        print(f"convert_json_to_excel fallback to csv: {e}")
        headers, projected_rows = (
            _project_internal_rows(rows)
            if include_internal_fields
            else _project_public_rows(rows)
        )
        df = pd.DataFrame(projected_rows, columns=headers)
        # 中文注释：使用 utf-8-sig，减少 Excel 打开 CSV 的乱码风险。
        return df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")

