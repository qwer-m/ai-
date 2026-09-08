from __future__ import annotations

import io
from zipfile import ZipFile

from openpyxl import load_workbook

from modules.agent_platform.excel_export import build_test_cases_excel


def test_export_uses_eight_chinese_columns_and_separates_step_results() -> None:
    excel_bytes = build_test_cases_excel(
        [
            {
                "case_id": "TC-001",
                "title": "提交学习计划",
                "module": "学习计划",
                "preconditions": ["用户已登录", "已有可用课程"],
                "test_input": "角色=管理员；项目名称=秋季学习计划",
                "steps": [
                    {"action": "进入计划页", "expected": "展示计划列表"},
                    {"action": "点击提交", "expected": "计划保存成功"},
                ],
                "priority": "P0",
            }
        ]
    )

    workbook = load_workbook(io.BytesIO(excel_bytes), data_only=False)
    worksheet = workbook["测试用例"]
    assert [cell.value for cell in worksheet[1]] == [
        "用例编号",
        "用例标题",
        "测试模块",
        "前置条件",
        "执行步骤",
        "测试输入",
        "预期结果",
        "用例级别",
    ]
    assert worksheet["D2"].value == "用户已登录\n已有可用课程"
    assert worksheet["E2"].value == "1. 进入计划页\n2. 点击提交"
    assert worksheet["F2"].value == "角色=管理员；项目名称=秋季学习计划"
    assert worksheet["G2"].value == "1. 展示计划列表\n2. 计划保存成功"
    assert worksheet.freeze_panes == "A2"
    assert worksheet.auto_filter.ref == "A1:H2"
    assert worksheet["A1"].fill.fgColor.rgb == "008EDB82"
    assert worksheet["A1"].font.color.rgb == "00111827"
    assert worksheet.row_dimensions[2].height >= 30

    with ZipFile(io.BytesIO(excel_bytes)) as archive:
        assert archive.testzip() is None


def test_export_writes_formula_like_content_as_text() -> None:
    excel_bytes = build_test_cases_excel(
        [
            {
                "case_id": "TC-002",
                "title": '=HYPERLINK("https://invalid.example")',
                "module": "安全校验",
                "preconditions": [],
                "steps": [{"action": "+cmd", "expected": "不执行公式"}],
                "priority": "P1",
            }
        ]
    )

    worksheet = load_workbook(io.BytesIO(excel_bytes), data_only=False)["测试用例"]
    assert worksheet["B2"].data_type == "s"
    assert worksheet["B2"].value.startswith("'")
    assert worksheet["E2"].data_type == "s"


def test_export_expands_row_height_for_wrapped_chinese_text() -> None:
    excel_bytes = build_test_cases_excel(
        [
            {
                "case_id": "TC-003",
                "title": "长文本展示",
                "module": "导出",
                "preconditions": [],
                "steps": [
                    {
                        "action": "查看导出的测试用例",
                        "expected": "导出的长文本能够根据列宽自动折行，并且不会被固定行高截断" * 4,
                    }
                ],
                "priority": "P2",
            }
        ]
    )

    worksheet = load_workbook(io.BytesIO(excel_bytes))["测试用例"]
    assert worksheet.row_dimensions[2].height > 30
