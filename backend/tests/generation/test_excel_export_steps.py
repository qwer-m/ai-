from openpyxl import load_workbook

from modules.test_generation_components.export.excel_export import convert_json_to_excel


def _read_first_steps_cell(rows):
    excel_bytes = convert_json_to_excel(rows)
    workbook = load_workbook(filename=__import__("io").BytesIO(excel_bytes))
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    steps_col = headers.index("测试步骤") + 1
    return sheet.cell(row=2, column=steps_col).value


def _read_sheet(rows, *, include_internal_fields=False):
    excel_bytes = convert_json_to_excel(rows, include_internal_fields=include_internal_fields)
    workbook = load_workbook(filename=__import__("io").BytesIO(excel_bytes))
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    values = [cell.value for cell in sheet[2]]
    return headers, dict(zip(headers, values))


def _read_first_column_values(rows):
    excel_bytes = convert_json_to_excel(rows)
    workbook = load_workbook(filename=__import__("io").BytesIO(excel_bytes))
    sheet = workbook.active
    return [sheet.cell(row=row_index, column=1).value for row_index in range(2, sheet.max_row + 1)]


def test_excel_export_does_not_duplicate_existing_step_numbers():
    value = _read_first_steps_cell(
        [
            {
                "id": "TC-001",
                "steps": ["1. open workbook", "2. find question card"],
            }
        ]
    )

    assert value == "1. open workbook\n2. find question card"
    assert "1. 1." not in value
    assert "2. 2." not in value


def test_excel_export_adds_numbers_for_plain_steps():
    value = _read_first_steps_cell(
        [
            {
                "id": "TC-002",
                "steps": ["open workbook", "check list"],
            }
        ]
    )

    assert value == "1. open workbook\n2. check list"


def test_excel_export_uses_presentation_order_when_available():
    ids = _read_first_column_values(
        [
            {"id": "TC-002", "description": "execution second", "steps": ["b"], "presentation_order": 2},
            {"id": "TC-001", "description": "read first", "steps": ["a"], "presentation_order": 1},
            {"id": "TC-003", "description": "fallback", "steps": ["c"]},
        ]
    )

    assert ids == ["TC-001", "TC-002", "TC-003"]


def test_excel_export_defaults_to_user_facing_chinese_columns():
    headers, row = _read_sheet(
        [
            {
                "id": "TC-001",
                "description": "create plan",
                "test_module": "schedule",
                "preconditions": ["logged in"],
                "steps": ["open planner"],
                "test_input": "valid plan",
                "expected_result": "plan is saved",
                "priority": "P0",
                "priority_final": "P0",
                "workflow_id": "schedule_flow",
                "source_state": "draft",
                "target_state": "saved",
                "execution_group": "main_smoke",
                "role": "student",
                "session_key": "student_session",
                "depends_on": ["TC-000"],
            }
        ]
    )

    assert headers == [
        "用例ID",
        "用例标题",
        "所属模块",
        "前置条件",
        "测试步骤",
        "测试数据",
        "预期结果",
        "优先级",
    ]
    assert "priority_final" not in headers
    assert "workflow_id" not in headers
    assert "execution_group" not in headers
    assert row["用例ID"] == "TC-001"
    assert row["用例标题"] == "create plan"
    assert row["所属模块"] == "schedule"
    assert row["前置条件"] == "logged in"
    assert row["测试步骤"] == "1. open planner"
    assert row["测试数据"] == "valid plan"
    assert row["预期结果"] == "plan is saved"
    assert row["优先级"] == "P0"


def test_excel_export_accepts_alias_case_fields():
    rows = [
        {
            "caseId": "TC-ALIAS",
            "title": "create plan from alias",
            "testModule": "schedule",
            "precondition": ["logged in"],
            "testSteps": [
                {"action": "open planner", "expect": "form visible"},
                {"action": "save", "expect": "toast visible"},
            ],
            "testInput": "valid alias plan",
            "expectedResult": "plan is saved",
            "Priority": "P2",
            "finalPriority": "P0",
        }
    ]

    headers, public_row = _read_sheet(rows)
    assert public_row[headers[0]] == "TC-ALIAS"
    assert public_row[headers[1]] == "create plan from alias"
    assert public_row[headers[2]] == "schedule"
    assert public_row[headers[3]] == "logged in"
    assert public_row[headers[4]] == "1. open planner form visible\n2. save toast visible"
    assert public_row[headers[5]] == "valid alias plan"
    assert public_row[headers[6]] == "plan is saved"
    assert public_row[headers[7]] == "P0"

    _headers, internal_row = _read_sheet(rows, include_internal_fields=True)
    assert internal_row["id"] == "TC-ALIAS"
    assert internal_row["description"] == "create plan from alias"
    assert internal_row["test_module"] == "schedule"
    assert internal_row["priority"] == "P2"
    assert internal_row["priority_final"] == "P0"


def test_excel_export_keeps_literal_list_preconditions_compatibility():
    rows = [
        {
            "id": "TC-LITERAL",
            "description": "literal preconditions",
            "test_module": "schedule",
            "preconditions": "['logged in', {'state': 'seeded'}]",
            "steps": ["open planner"],
            "expected_result": "planner opens",
            "priority": "P1",
        }
    ]

    headers, public_row = _read_sheet(rows)

    assert public_row[headers[3]] == "logged in\nseeded"


def test_excel_export_internal_mode_includes_persistable_execution_fields():
    headers, row = _read_sheet(
        [
            {
                "id": "TC-001",
                "description": "create plan",
                "test_module": "schedule",
                "preconditions": ["logged in"],
                "steps": ["open planner"],
                "test_input": "valid plan",
                "expected_result": "plan is saved",
                "priority": "P1",
                "priority_final": "P0",
                "workflow_id": "schedule_flow",
                "source_state": "draft",
                "target_state": "saved",
                "execution_group": "main_smoke",
                "role": "student",
                "session_key": "student_session",
                "depends_on": ["TC-000"],
            }
        ],
        include_internal_fields=True,
    )

    assert headers[:8] == [
        "id",
        "description",
        "test_module",
        "preconditions",
        "steps",
        "test_input",
        "expected_result",
        "priority",
    ]
    assert "priority_final" in headers
    assert "workflow_id" in headers
    assert "execution_group" in headers
    assert row["priority_final"] == "P0"
    assert row["workflow_id"] == "schedule_flow"
    assert row["source_state"] == "draft"
    assert row["target_state"] == "saved"
    assert row["execution_group"] == "main_smoke"
    assert row["role"] == "student"
    assert row["session_key"] == "student_session"
    assert row["depends_on"] == "TC-000"


def test_excel_export_includes_execution_suite_sheets_by_default():
    excel_bytes = convert_json_to_excel(
        [
            {
                "id": "TC-001",
                "description": "创建学习计划",
                "test_module": "学习计划",
                "steps": ["打开学习计划页", "填写计划", "提交保存"],
                "expected_result": "计划保存成功",
                "priority": "P0",
                "execution_group": "main_smoke",
                "execution_sequence": 1,
                "role": "teacher",
                "session_key": "teacher_session",
            },
            {
                "id": "TC-002",
                "description": "学生查看学习计划",
                "test_module": "学习计划",
                "steps": ["切换学生账号", "打开学习计划"],
                "expected_result": "可以看到刚保存的计划",
                "priority": "P0",
                "execution_group": "main_smoke",
                "execution_sequence": 2,
                "role": "student",
                "session_key": "student_session",
                "depends_on": ["TC-001"],
            },
        ]
    )

    workbook = load_workbook(filename=__import__("io").BytesIO(excel_bytes))
    assert workbook.sheetnames == ["测试用例", "执行套件", "执行顺序"]

    run_sheet = workbook["执行顺序"]
    headers = [cell.value for cell in run_sheet[1]]
    case_id_col = headers.index("用例ID") + 1
    role_col = headers.index("角色") + 1
    depends_col = headers.index("依赖用例") + 1

    assert run_sheet.cell(row=2, column=case_id_col).value == "TC-001"
    assert run_sheet.cell(row=3, column=case_id_col).value == "TC-002"
    assert run_sheet.cell(row=3, column=role_col).value == "student"
    assert run_sheet.cell(row=3, column=depends_col).value == "TC-001"
