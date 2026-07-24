from __future__ import annotations

from openpyxl import load_workbook

from modules.test_generation_components.execution.execution_suite import (
    build_execution_suite,
    convert_execution_suite_to_excel,
    parse_generated_cases_payload,
)


def _sample_cases() -> list[dict]:
    return [
        {
            "id": "TC-001",
            "description": "create schedule plan",
            "test_module": "schedule",
            "preconditions": ["teacher logged in"],
            "steps": ["open scheduler", "save plan"],
            "test_input": "valid course and time",
            "expected_result": "plan is saved",
            "priority": "P0",
            "priority_final": "P0",
            "execution_group": "main_smoke",
            "execution_sequence": 1,
            "chain_id": "main_smoke_chain",
            "role": "teacher",
            "session_key": "teacher_session",
            "fixture_key": "workflow_blueprint_chain_seed",
            "group_setup": "seed_workflow_blueprint_dataset()",
            "group_teardown": "cleanup_workflow_blueprint_dataset()",
            "setup_hint": "run group setup first",
            "teardown_hint": "cleanup at the end",
        },
        {
            "id": "TC-002",
            "description": "student sees saved plan",
            "test_module": "student home",
            "preconditions": ["plan has been saved"],
            "steps": ["open student home"],
            "test_input": "student account",
            "expected_result": "saved plan is visible",
            "priority": "P0",
            "priority_final": "P0",
            "execution_group": "main_smoke",
            "execution_sequence": 2,
            "chain_id": "main_smoke_chain",
            "depends_on": ["TC-001"],
            "role": "student",
            "session_key": "student_session",
            "fixture_key": "workflow_blueprint_chain_seed",
            "group_setup": "seed_workflow_blueprint_dataset()",
            "group_teardown": "cleanup_workflow_blueprint_dataset()",
            "setup_hint": "reuse TC-001 result",
            "teardown_hint": "cleanup at the end",
        },
        {
            "id": "TC-003",
            "description": "unauthorized user cannot open schedule",
            "test_module": "permission",
            "preconditions": ["guest user"],
            "steps": ["open schedule page"],
            "test_input": "guest account",
            "expected_result": "access is denied",
            "priority": "P1",
            "priority_final": "P1",
            "execution_group": "permission",
            "execution_sequence": 3,
            "role": "guest",
            "session_key": "guest_session",
            "fixture_key": "permission_state_dataset",
            "setup_hint": "seed guest account",
            "teardown_hint": "reset permissions",
        },
    ]


def test_execution_suite_groups_cases_into_runnable_suites() -> None:
    suite = build_execution_suite(_sample_cases())

    assert suite["kind"] == "execution_suite"
    assert suite["case_count"] == 3
    assert suite["suite_count"] == 2
    assert suite["linear_executable"] is True
    assert suite["execution_readiness"] == "ready"
    assert suite["warnings"] == []
    assert suite["metadata_quality"]["complete_execution_metadata"] is True
    assert suite["flat_run_order"][0]["case_id"] == "TC-001"
    assert suite["flat_run_order"][1]["depends_on"] == ["TC-001"]

    main_suite = suite["suites"][0]
    assert main_suite["suite_id"] == "main_smoke_chain"
    assert main_suite["suite_name"] == "主链路冒烟"
    assert main_suite["run_mode"] == "sequential"
    assert main_suite["runnable"] is True
    assert main_suite["roles"] == ["student", "teacher"]
    assert [case["case_id"] for case in main_suite["cases"]] == ["TC-001", "TC-002"]

    permission_suite = suite["suites"][1]
    assert permission_suite["suite_id"] == "permission_suite"
    assert permission_suite["run_mode"] == "isolated"
    assert permission_suite["runnable"] is True


def test_execution_suite_uses_alias_fields_for_case_payload() -> None:
    cases = [
        {
            "caseId": "TC-ALIAS",
            "title": "save schedule plan",
            "testModule": "schedule",
            "precondition": ["teacher logged in"],
            "testSteps": ["open scheduler", "save plan"],
            "testInput": "valid course and time",
            "expectedResult": "plan is saved",
            "priorityFinal": "P0",
            "execution_group": "main_smoke",
            "execution_sequence": 1,
            "role": "teacher",
            "session_key": "teacher_session",
        }
    ]

    suite = build_execution_suite(cases)
    case_ref = suite["suites"][0]["cases"][0]

    assert case_ref["case_id"] == "TC-ALIAS"
    assert case_ref["description"] == "save schedule plan"
    assert case_ref["test_module"] == "schedule"
    assert case_ref["preconditions"] == ["teacher logged in"]
    assert case_ref["steps"] == ["open scheduler", "save plan"]
    assert case_ref["test_input"] == "valid course and time"
    assert case_ref["expected_result"] == "plan is saved"
    assert case_ref["priority"] == "P0"
    assert case_ref["action"] == "open scheduler"
    assert case_ref["runnable"] is True


def test_execution_suite_projects_public_step_as_display_action() -> None:
    cases = [
        {
            "id": "TC-ACTION",
            "description": "submit order",
            "steps": ["click submit", "verify success"],
            "expected_result": "order is submitted",
            "action": "submit order from checkout",
            "execution_group": "main_smoke",
            "execution_sequence": 1,
            "role": "buyer",
            "session_key": "buyer_session",
        }
    ]

    suite = build_execution_suite(cases)
    case_ref = suite["suites"][0]["cases"][0]

    assert case_ref["action"] == "click submit"
    assert case_ref["transition_action"] == "submit order from checkout"


def test_execution_suite_falls_back_to_explicit_action_without_steps() -> None:
    cases = [
        {
            "id": "TC-ACTION",
            "description": "submit order",
            "expected_result": "order is submitted",
            "action": "submit order from checkout",
            "execution_group": "main_smoke",
            "execution_sequence": 1,
            "role": "buyer",
            "session_key": "buyer_session",
        }
    ]

    suite = build_execution_suite(cases)
    case_ref = suite["suites"][0]["cases"][0]

    assert case_ref["action"] == "submit order from checkout"
    assert case_ref["transition_action"] == "submit order from checkout"
    assert case_ref["runnable"] is False


def test_execution_suite_preserves_transition_action_when_present() -> None:
    cases = [
        {
            "id": "TC-ACTION",
            "description": "submit order",
            "steps": ["click submit", "verify success"],
            "expected_result": "order is submitted",
            "action": "click submit",
            "transition_action": "submit order from checkout",
            "execution_group": "main_smoke",
            "execution_sequence": 1,
            "role": "buyer",
            "session_key": "buyer_session",
        }
    ]

    suite = build_execution_suite(cases)
    case_ref = suite["suites"][0]["cases"][0]

    assert case_ref["action"] == "click submit"
    assert case_ref["transition_action"] == "submit order from checkout"


def test_execution_suite_splits_text_steps_with_shared_accessor() -> None:
    cases = [
        {
            "caseId": "TC-TEXT-STEPS",
            "title": "save schedule plan",
            "testModule": "schedule",
            "testSteps": "open scheduler;save plan；verify toast",
            "expectedResult": "plan is saved",
            "execution_group": "main_smoke",
            "execution_sequence": 1,
            "role": "teacher",
            "session_key": "teacher_session",
        }
    ]

    suite = build_execution_suite(cases)
    case_ref = suite["suites"][0]["cases"][0]

    assert case_ref["steps"] == ["open scheduler", "save plan", "verify toast"]
    assert case_ref["runnable"] is True


def test_execution_suite_splits_text_preconditions_and_dependencies() -> None:
    cases = [
        {
            "id": "TC-001",
            "description": "seed schedule plan",
            "steps": ["seed plan"],
            "expected_result": "plan exists",
            "execution_group": "main_smoke",
            "execution_sequence": 1,
            "role": "teacher",
            "session_key": "teacher_session",
        },
        {
            "id": "TC-002",
            "description": "save schedule plan",
            "precondition": "teacher logged in\nplan draft exists",
            "steps": ["save plan"],
            "expected_result": "plan is saved",
            "execution_group": "main_smoke",
            "execution_sequence": 2,
            "role": "teacher",
            "session_key": "teacher_session",
            "depends_on": "TC-001\n",
        },
    ]

    suite = build_execution_suite(cases)
    case_ref = suite["suites"][0]["cases"][1]

    assert case_ref["preconditions"] == ["teacher logged in", "plan draft exists"]
    assert case_ref["depends_on"] == ["TC-001"]
    assert suite["linear_executable"] is True


def test_execution_suite_marks_missing_dependencies() -> None:
    cases = _sample_cases()
    cases[1]["depends_on"] = ["TC-404"]

    suite = build_execution_suite(cases)
    main_suite = suite["suites"][0]

    assert suite["linear_executable"] is False
    assert suite["execution_readiness"] == "partial"
    assert "存在 1 个缺失依赖引用" in suite["warnings"]
    assert main_suite["runnable"] is False
    assert main_suite["missing_dependencies"] == ["TC-404"]
    assert "存在缺失依赖用例" in main_suite["warnings"][0]


def test_execution_suite_marks_legacy_payload_as_manual_sequence() -> None:
    cases = [
        {
            "id": "TC-001",
            "description": "open legacy page",
            "steps": ["open page"],
            "expected_result": "page is visible",
            "priority_final": "P1",
        },
        {
            "id": "TC-002",
            "description": "submit legacy form",
            "steps": ["submit form"],
            "expected_result": "form is saved",
            "priority_final": "P1",
        },
    ]

    suite = build_execution_suite(cases)

    assert suite["linear_executable"] is False
    assert suite["execution_readiness"] == "legacy_manual"
    assert suite["metadata_quality"]["has_any_execution_metadata"] is False
    assert "历史结果缺少执行元数据" in suite["warnings"][0]
    assert "缺少 main_smoke 主链" in suite["warnings"][1]
    assert suite["suites"][0]["suite_id"] == "unknown_suite"
    assert suite["suites"][0]["suite_name"] == "手工执行顺序"
    assert suite["suites"][0]["run_mode"] == "manual_sequential"
    assert [item["case_id"] for item in suite["suites"][0]["cases"]] == ["TC-001", "TC-002"]


def test_execution_suite_without_main_smoke_reports_partial_readiness() -> None:
    cases = _sample_cases()
    for case in cases:
        if case["execution_group"] == "main_smoke":
            case["execution_group"] = "independent_functional"

    suite = build_execution_suite(cases)

    assert suite["linear_executable"] is False
    assert suite["execution_readiness"] == "partial"
    assert "缺少 main_smoke 主链" in suite["warnings"][0]
    assert suite["main_suite_id"] == ""


def test_execution_suite_marks_declared_independent_suite_ready_without_main_chain_warning() -> None:
    cases = _sample_cases()
    for case in cases:
        if case["execution_group"] == "main_smoke":
            case["execution_group"] = "independent_functional"

    suite = build_execution_suite(cases, workflow_absence_declared=True)

    assert suite["linear_executable"] is False
    assert suite["workflow_absence_declared"] is True
    assert suite["execution_readiness"] == "independent_ready"
    assert all("main_smoke" not in warning for warning in suite["warnings"])
    assert suite["main_suite_id"] == ""


def test_execution_suite_orders_side_suites_with_shared_execution_rank() -> None:
    cases = [
        {
            "id": "TC-display",
            "description": "display order status",
            "steps": ["open order detail"],
            "expected_result": "status is displayed",
            "execution_group": "display",
            "execution_sequence": 6,
            "role": "viewer",
            "session_key": "viewer_session",
        },
        {
            "id": "TC-independent",
            "description": "recalculate coupon",
            "steps": ["apply coupon"],
            "expected_result": "total is recalculated",
            "execution_group": "independent",
            "execution_sequence": 5,
            "role": "student",
            "session_key": "student_session",
        },
        {
            "id": "TC-permission",
            "description": "block readonly submit",
            "steps": ["submit as readonly"],
            "expected_result": "permission denied",
            "execution_group": "permission",
            "execution_sequence": 3,
            "role": "readonly",
            "session_key": "readonly_session",
        },
    ]

    suite = build_execution_suite(cases)

    assert [item["execution_group"] for item in suite["suites"]] == [
        "permission",
        "independent",
        "display",
    ]
    assert [item["case_id"] for item in suite["flat_run_order"]] == [
        "TC-permission",
        "TC-independent",
        "TC-display",
    ]


def test_parse_generated_cases_payload_accepts_persisted_shapes() -> None:
    cases = _sample_cases()

    assert parse_generated_cases_payload(cases) == cases
    assert parse_generated_cases_payload({"cases": cases}) == cases
    assert parse_generated_cases_payload({"generated_result": cases}) == cases


def test_execution_suite_excel_contains_summary_and_run_order() -> None:
    excel_bytes = convert_execution_suite_to_excel(build_execution_suite(_sample_cases()))
    workbook = load_workbook(filename=__import__("io").BytesIO(excel_bytes))

    assert workbook.sheetnames == ["执行套件", "执行顺序"]
    summary = workbook["执行套件"]
    run_order = workbook["执行顺序"]

    assert summary.cell(row=1, column=1).value == "套件ID"
    assert summary.cell(row=2, column=1).value == "main_smoke_chain"
    assert summary.cell(row=2, column=10).value == "是"
    assert summary.cell(row=1, column=12).value == "诊断提示"
    assert run_order.cell(row=1, column=5).value == "用例ID"
    assert run_order.cell(row=3, column=5).value == "TC-002"
    assert run_order.cell(row=3, column=11).value == "TC-001"
