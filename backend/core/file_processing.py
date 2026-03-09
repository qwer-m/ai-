"""
文件处理模块 (File Processing Module)

该模块负责解析上传的文件内容，支持多种格式。
主要功能：
1. PDF 解析: 使用 pypdf 提取文本。
2. CSV 解析: 使用 pandas 转换为 CSV 字符串。
3. 图片处理: 转换为 Base64 编码 (OCR 预处理)。
4. 文本文件: 直接解码 UTF-8 内容。

被调用方：
- modules.knowledge_base (上传文档时解析内容)
"""

import io
import pypdf
import pandas as pd
from fastapi import UploadFile
import base64
from core.ai_client import ai_client
from typing import Optional

async def parse_file_content(file: UploadFile, image_prompt: str = "OCR: Extract all text from this image.") -> str:
    """
    解析文件内容 (Parse File Content)
    
    根据文件扩展名自动选择解析策略。
    
    Args:
        file: FastAPI UploadFile 对象。
        image_prompt: 图片 OCR 的提示词 (当前暂未完全实现 OCR 逻辑，仅占位)。
        
    Returns:
        str: 解析后的文本内容。
    """
    filename = file.filename.lower()
    content_bytes = await file.read()
    text_content = ""

    try:
        if filename.endswith(".pdf"):
            # Handle PDF
            pdf_file = io.BytesIO(content_bytes)
            reader = pypdf.PdfReader(pdf_file)
            for page in reader.pages:
                text_content += page.extract_text() + "\n"
        
        elif filename.endswith(('.xls', '.xlsx')):
            # Handle Excel with openpyxl for better formatting (merged cells)
            import openpyxl
            excel_file = io.BytesIO(content_bytes)
            try:
                wb = openpyxl.load_workbook(excel_file, data_only=True)
                text_content = ""
                
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    text_content += f"<h5>Sheet: {sheet_name}</h5>"
                    # Add custom CSS for Excel-like look
                    text_content += '<div class="table-responsive mb-4"><table class="table table-bordered table-sm table-hover" style="border-collapse: collapse; min-width: 100%; font-size: 0.9em;">'
                    
                    # Pre-calculate merged cells map
                    # Key: (row, col) 1-based index
                    # Value: (rowspan, colspan) or 'skip'
                    merge_map = {}
                    for merge_range in ws.merged_cells.ranges:
                        min_col, min_row, max_col, max_row = merge_range.min_col, merge_range.min_row, merge_range.max_col, merge_range.max_row
                        # Mark top-left cell with span info
                        merge_map[(min_row, min_col)] = (max_row - min_row + 1, max_col - min_col + 1)
                        # Mark all other cells in range as 'skip'
                        for r in range(min_row, max_row + 1):
                            for c in range(min_col, max_col + 1):
                                if r == min_row and c == min_col:
                                    continue
                                merge_map[(r, c)] = 'skip'

                    # Iterate rows
                    for r, row in enumerate(ws.iter_rows(), start=1):
                        text_content += "<tr>"
                        for c, cell in enumerate(row, start=1):
                            # Check merge map
                            if (r, c) in merge_map:
                                if merge_map[(r, c)] == 'skip':
                                    continue
                                rowspan, colspan = merge_map[(r, c)]
                                val = str(cell.value) if cell.value is not None else ""
                                # Style for merged cells
                                style = 'vertical-align: middle; white-space: pre-wrap;'
                                if rowspan > 1 or colspan > 1:
                                    style += ' background-color: #f8f9fa; font-weight: 500;'
                                text_content += f'<td rowspan="{rowspan}" colspan="{colspan}" style="{style}">{val}</td>'
                            else:
                                val = str(cell.value) if cell.value is not None else ""
                                text_content += f'<td style="white-space: pre-wrap;">{val}</td>'
                        text_content += "</tr>"
                    text_content += "</table></div>"
            except Exception as e:
                text_content = f"[Error reading Excel: {str(e)}]"
        
        elif filename.endswith(".csv"):
            # Handle CSV
            csv_file = io.BytesIO(content_bytes)
            try:
                df = pd.read_csv(csv_file)
                text_content = df.to_csv(index=False)
            except Exception as e:
                 # Fallback
                 try:
                    text_content = content_bytes.decode("utf-8")
                 except:
                    text_content = f"[Error reading CSV: {str(e)}]"
        
        elif filename.endswith(('.png', '.jpg', '.jpeg', '.bmp', '.gif', '.webp')):
             # Handle Image (OCR)
             try:
                 encoded_image = base64.b64encode(content_bytes).decode('utf-8')
                 # Placeholder for OCR
                 text_content = f"[Image Content: {filename}]"
             except Exception as e:
                 text_content = f"[Error processing image: {str(e)}]"
        
        else:
            # Try text
            try:
                text_content = content_bytes.decode("utf-8")
            except:
                text_content = f"[Unsupported file type: {filename}]"

    except Exception as e:
        text_content = f"[Error parsing file: {str(e)}]"

    return text_content