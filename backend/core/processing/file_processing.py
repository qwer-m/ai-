"""File parsing helpers.

Provides upload, byte, and path based parsing for documents and images,
including local/cloud OCR fallback behavior for image inputs.
"""

from __future__ import annotations

import io
import os
import tempfile
from html import escape
from pathlib import Path
from typing import Any, Optional

import pandas as pd
import pypdf
from fastapi import UploadFile
from sqlalchemy.orm import Session

IMAGE_EXTENSIONS = (".png", ".jpg", ".jpeg", ".bmp", ".gif", ".webp")
def _escape_preview_html_value(value: Any) -> str:
    if value is None:
        return ""
    return escape(str(value), quote=True)


def is_image_filename(filename: str) -> bool:
    return (filename or "").lower().endswith(IMAGE_EXTENSIONS)


def is_pdf_filename(filename: str) -> bool:
    return (filename or "").lower().endswith(".pdf")


def _is_ocr_failure_text(text: str) -> bool:
    value = (text or "").strip()
    if not value:
        return True

    lowered = value.lower()
    if lowered.startswith(("ocr error", "ocr exception", "error:", "exception", "[image ocr failed:", "[error processing image:")):
        return True

    failure_markers = (
        "额度耗尽",
        "免费额度已用完",
        "余额不足",
        "insufficient_quota",
        "quota exceeded",
        "rate limit",
        "authentication failed",
        "invalid api key",
        "model not found",
    )
    value_lower = value.lower()
    return any(marker in value or marker in value_lower for marker in failure_markers)


def _is_meaningful_ocr_text(text: str) -> bool:
    value = (text or "").strip()
    if not value:
        return False
    if _is_ocr_failure_text(value):
        return False
    non_space = "".join(ch for ch in value if not ch.isspace())
    return len(non_space) >= 6


def _normalize_tesseract_cmd(raw_cmd: Optional[str]) -> str:
    if not raw_cmd:
        return ""
    value = str(raw_cmd).strip().strip('"').strip("'")
    if not value:
        return ""
    return os.path.expandvars(os.path.expanduser(value))


def _resolve_tesseract_cmd(db: Optional[Session], user_id: Optional[int]) -> str:
    if db is None or not user_id:
        return ""
    try:
        from core.settings.config_manager import config_manager

        active_config = config_manager.get_active_config(db, user_id)
        if not active_config:
            return ""
        metadata = active_config.metadata_info if isinstance(active_config.metadata_info, dict) else {}
        return _normalize_tesseract_cmd(metadata.get("tesseract_path"))
    except Exception:
        return ""


def _run_local_ocr(content_bytes: bytes, tesseract_cmd: Optional[str] = None) -> tuple[str, str]:
    try:
        from PIL import Image, ImageOps
        import pytesseract
    except Exception as e:
        return "", f"local_ocr_dependency_missing: {e}"

    try:
        configured_cmd = _normalize_tesseract_cmd(tesseract_cmd)
        if configured_cmd:
            pytesseract.pytesseract.tesseract_cmd = configured_cmd

        image = Image.open(io.BytesIO(content_bytes))
        image = ImageOps.grayscale(image)
        text = pytesseract.image_to_string(image, lang="chi_sim+eng")
        if _is_meaningful_ocr_text(text):
            return text.strip(), ""
        text_eng = pytesseract.image_to_string(image, lang="eng")
        if _is_meaningful_ocr_text(text_eng):
            return text_eng.strip(), ""
        return "", "local_ocr_empty"
    except Exception as e:
        return "", f"local_ocr_error: {e}"


def parse_image_bytes_with_fallback(
    filename: str,
    content_bytes: bytes,
    image_prompt: str = "OCR: Extract all text from this image.",
    db: Optional[Session] = None,
    user_id: Optional[int] = None,
) -> tuple[str, dict[str, Any]]:
    meta: dict[str, Any] = {
        "ocr_source": "none",
        "local_ocr_used": False,
        "local_ocr_error": "",
        "cloud_ocr_used": False,
        "cloud_fallback": False,
        "error": "",
    }

    tesseract_cmd = _resolve_tesseract_cmd(db, user_id)
    local_text, local_error = _run_local_ocr(content_bytes, tesseract_cmd=tesseract_cmd)
    meta["local_ocr_error"] = local_error
    if _is_meaningful_ocr_text(local_text):
        meta["ocr_source"] = "local"
        meta["local_ocr_used"] = True
        return local_text.strip(), meta

    if db is None or not user_id:
        meta["ocr_source"] = "offline_fallback"
        meta["error"] = local_error or "local_ocr_unavailable_or_empty"
        return f"[Image Content: {(filename or '').lower()}]", meta

    from core.ai.ai_client import get_client_for_user

    suffix = Path((filename or "").lower()).suffix or ".png"
    temp_path = ""
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as temp_file:
            temp_file.write(content_bytes)
            temp_path = temp_file.name

        client = get_client_for_user(user_id, db)
        ocr_result = client.analyze_image(
            f"file://{temp_path}",
            prompt=image_prompt,
            db=db,
        )
        meta["cloud_ocr_used"] = True
        meta["cloud_fallback"] = True

        if isinstance(ocr_result, str):
            ocr_text = ocr_result.strip()
            if _is_meaningful_ocr_text(ocr_text):
                meta["ocr_source"] = "cloud"
                return ocr_text, meta
            meta["ocr_source"] = "failed"
            meta["error"] = ocr_text or "empty response"
            return f"[Image OCR Failed: {ocr_text or 'empty response'}]", meta

        meta["ocr_source"] = "failed"
        meta["error"] = "invalid response"
        return "[Image OCR Failed: invalid response]", meta
    except Exception as e:
        meta["ocr_source"] = "failed"
        meta["error"] = str(e)
        return f"[Error processing image: {str(e)}]", meta
    finally:
        if temp_path and os.path.exists(temp_path):
            os.remove(temp_path)


def _parse_pdf_text_with_meta(content_bytes: bytes) -> tuple[str, dict[str, Any]]:
    meta: dict[str, Any] = {
        "pdf_page_count": 0,
        "pdf_text_page_count": 0,
        "pdf_text_chars": 0,
        "pdf_text_error": "",
    }
    text_parts: list[str] = []
    pdf_file = io.BytesIO(content_bytes)
    reader = pypdf.PdfReader(pdf_file)
    meta["pdf_page_count"] = int(len(reader.pages))
    for page in reader.pages:
        page_text = page.extract_text() or ""
        if page_text.strip():
            meta["pdf_text_page_count"] += 1
        text_parts.append(page_text)
    text_content = "\n".join(text_parts).strip()
    meta["pdf_text_chars"] = len(text_content)
    return text_content + ("\n" if text_content else ""), meta


def parse_file_bytes(
    filename: str,
    content_bytes: bytes,
    image_prompt: str = "OCR: Extract all text from this image.",
    db: Optional[Session] = None,
    user_id: Optional[int] = None,
) -> str:
    """Parse uploaded file bytes into text or HTML preview content."""
    lowered_name = (filename or "").lower()
    text_content = ""

    try:
        if is_pdf_filename(lowered_name):
            text_content, _pdf_meta = _parse_pdf_text_with_meta(content_bytes)

        elif lowered_name.endswith((".xls", ".xlsx")):
            import openpyxl

            excel_file = io.BytesIO(content_bytes)
            try:
                wb = openpyxl.load_workbook(excel_file, data_only=True)
                text_content = ""
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    safe_sheet_name = _escape_preview_html_value(sheet_name)
                    text_content += f"<h5>Sheet: {safe_sheet_name}</h5>"
                    text_content += (
                        '<div class="table-responsive mb-4"><table class="table table-bordered '
                        'table-sm table-hover" style="border-collapse: collapse; min-width: 100%; '
                        'font-size: 0.9em;">'
                    )

                    merge_map: dict[tuple[int, int], tuple[int, int] | str] = {}
                    for merge_range in ws.merged_cells.ranges:
                        min_col, min_row = merge_range.min_col, merge_range.min_row
                        max_col, max_row = merge_range.max_col, merge_range.max_row
                        merge_map[(min_row, min_col)] = (
                            max_row - min_row + 1,
                            max_col - min_col + 1,
                        )
                        for row in range(min_row, max_row + 1):
                            for col in range(min_col, max_col + 1):
                                if row == min_row and col == min_col:
                                    continue
                                merge_map[(row, col)] = "skip"

                    for row_idx, row in enumerate(ws.iter_rows(), start=1):
                        text_content += "<tr>"
                        for col_idx, cell in enumerate(row, start=1):
                            if (row_idx, col_idx) in merge_map:
                                if merge_map[(row_idx, col_idx)] == "skip":
                                    continue
                                rowspan, colspan = merge_map[(row_idx, col_idx)]
                                value = _escape_preview_html_value(cell.value)
                                style = "vertical-align: middle; white-space: pre-wrap;"
                                if rowspan > 1 or colspan > 1:
                                    style += " background-color: #f8f9fa; font-weight: 500;"
                                text_content += (
                                    f'<td rowspan="{rowspan}" colspan="{colspan}" '
                                    f'style="{style}">{value}</td>'
                                )
                            else:
                                value = _escape_preview_html_value(cell.value)
                                text_content += f'<td style="white-space: pre-wrap;">{value}</td>'
                        text_content += "</tr>"
                    text_content += "</table></div>"
            except Exception as e:
                text_content = f"[Error reading Excel: {str(e)}]"

        elif lowered_name.endswith(".csv"):
            csv_file = io.BytesIO(content_bytes)
            try:
                df = pd.read_csv(csv_file)
                text_content = df.to_csv(index=False)
            except Exception as e:
                try:
                    text_content = content_bytes.decode("utf-8")
                except Exception:
                    text_content = f"[Error reading CSV: {str(e)}]"

        elif is_image_filename(lowered_name):
            text_content, _meta = parse_image_bytes_with_fallback(
                filename=lowered_name,
                content_bytes=content_bytes,
                image_prompt=image_prompt,
                db=db,
                user_id=user_id,
            )

        else:
            try:
                text_content = content_bytes.decode("utf-8")
            except Exception:
                text_content = f"[Unsupported file type: {lowered_name}]"
    except Exception as e:
        text_content = f"[Error parsing file: {str(e)}]"

    return text_content


async def parse_file_content_with_meta(
    file: UploadFile,
    image_prompt: str = "OCR: Extract all text from this image.",
    db: Optional[Session] = None,
    user_id: Optional[int] = None,
) -> tuple[str, dict[str, Any]]:
    """Parse UploadFile and return lightweight source/OCR metadata."""
    filename = file.filename or ""
    content_bytes = await file.read()
    meta: dict[str, Any] = {
        "filename": filename,
        "content_type": file.content_type or "",
        "size": len(content_bytes),
        "is_image": is_image_filename(filename),
        "parse_strategy": "file_text",
    }
    if is_image_filename(filename):
        text_content, ocr_meta = parse_image_bytes_with_fallback(
            filename=filename,
            content_bytes=content_bytes,
            image_prompt=image_prompt,
            db=db,
            user_id=user_id,
        )
        meta.update(
            {
                "parse_strategy": "image_ocr",
                "ocr": dict(ocr_meta or {}),
            }
        )
        return text_content, meta

    if is_pdf_filename(filename):
        try:
            text_content, pdf_meta = _parse_pdf_text_with_meta(content_bytes)
            meta.update(pdf_meta)
            return text_content, meta
        except Exception as e:
            meta["pdf_text_error"] = str(e)

    text_content = parse_file_bytes(
        filename,
        content_bytes,
        image_prompt=image_prompt,
        db=db,
        user_id=user_id,
    )
    return text_content, meta


def parse_file_path(
    file_path: str,
    image_prompt: str = "OCR: Extract all text from this image.",
    db: Optional[Session] = None,
    user_id: Optional[int] = None,
) -> str:
    """Parse a local file path for offline tasks."""
    path = Path(file_path)
    content_bytes = path.read_bytes()
    return parse_file_bytes(
        path.name,
        content_bytes,
        image_prompt=image_prompt,
        db=db,
        user_id=user_id,
    )
