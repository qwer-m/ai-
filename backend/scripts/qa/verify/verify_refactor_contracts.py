"""只读回放真实运行，并对副本注入缺口验证契约；不调用模型或改写历史数据。"""

from __future__ import annotations

import argparse
from copy import deepcopy
from io import BytesIO
import json
from pathlib import Path
import sys

sys.path.insert(0, str(Path(__file__).resolve().parents[3]))

from openpyxl import load_workbook
from sqlalchemy import text

from core.db.database import SessionLocal
from core.db.model_defs import AgentNodeRun, AgentRun
from modules.agent_platform.excel_export import build_test_cases_excel, EXPORT_COLUMNS
from modules.agent_platform.results import persisted_test_generation_result
from modules.agent_platform.serialization import serialize_run
from modules.agent_platform.service import _normalize_workflow_input
from modules.agent_platform.test_generation_batching import (
    _validate_generation_batch_output,
    postprocess_generation_batch_item,
)
from modules.agent_platform.test_generation_facts import validate_case_fact_bindings
from modules.agent_platform.test_generation_review import _inline_repair_cases


def check_missing_fact(source: dict, output: dict) -> str | None:
    """只移除副本中的事实引用，检验审计不会暗中修补输入。"""
    inline = _inline_repair_cases(
        test_cases=output["test_cases"], bindings=output["case_fact_bindings"],
    )
    for fact_id in source["case_fact_contract"]["required_fact_ids"]:
        candidate = deepcopy(inline)
        groups = []
        for case in candidate:
            groups.extend(item["fact_ids"] for item in case["preconditions"])
            groups.extend(step["fact_bindings"]["expected"] for step in case["steps"])
        # 避免先触发非空断言，保证此次验证确实到达覆盖缺口检查。
        if any(group == [fact_id] for group in groups):
            continue
        for case in candidate:
            for step in case["steps"]:
                groups.append(step["fact_bindings"]["action"])
            groups.append(case["test_input"]["fact_ids"])
        for group in groups:
            group[:] = [value for value in group if value != fact_id]
        for case in candidate:
            if not case["test_input"]["fact_ids"]:
                case["test_input"]["text"] = ""
        arguments = {"item_input": deepcopy(source), "item_output": {"test_cases": candidate}}
        before = deepcopy(arguments)
        try:
            postprocess_generation_batch_item(None, arguments)
        except ValueError as exc:
            assert "未完整覆盖平台要求的事实" in str(exc), str(exc)
            assert fact_id in str(exc)
            assert arguments == before, "校验不得补写预期或事实绑定"
            return fact_id
        raise AssertionError(f"缺失事实引用仍通过审计: {fact_id}")
    return None


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--run-id", type=int, required=True)
    options = parser.parse_args()
    with SessionLocal() as db:
        db.execute(text("SET TRANSACTION READ ONLY"))
        run = db.get(AgentRun, options.run_id)
        if run is None or run.status != "success":
            raise ValueError("需要数据库中真实成功的运行记录")
        original_context = deepcopy(run.run_context)
        original_output = deepcopy(run.output_payload)
        artifact = persisted_test_generation_result(run)
        assert artifact is not None
        assert serialize_run(run)["test_generation_result"] == artifact
        cases = artifact["test_cases"]
        workbook = load_workbook(BytesIO(build_test_cases_excel(cases)))
        sheet = workbook.active
        assert [cell.value for cell in sheet[1]] == [name for name, _ in EXPORT_COLUMNS]
        assert sheet.max_row == len(cases) + 1
        input_column = next(i for i, (_, key) in enumerate(EXPORT_COLUMNS, 1) if key == "test_input")
        for row, case in enumerate(cases, 2):
            assert (sheet.cell(row, input_column).value or "") == case["test_input"]
        workbook.close()

        nodes = {
            node.node_key: node for node in sorted(db.query(AgentNodeRun).filter(
                AgentNodeRun.run_id == run.id,
                AgentNodeRun.node_key.in_(["generation", "prepare_generation"]),
            ).all(), key=lambda node: (node.attempt, node.id))
        }
        inputs = nodes["prepare_generation"].output_payload["items"]
        records = nodes["generation"].output_payload["items"]
        rejected_fact = None
        for record in records:
            source = inputs[record["item_index"]]
            output = record["output"]
            assert _validate_generation_batch_output(source_input=source, output=output) == output
            if rejected_fact is None:
                rejected_fact = check_missing_fact(source, output)
        assert rejected_fact is not None, "真实批次未包含可独立注入的事实绑定"

        first = deepcopy(cases[0])
        binding = deepcopy(next(item for item in artifact["case_fact_bindings"] if item["case_id"] == first["case_id"]))
        facts = run.run_context["artifacts"]["source_semantics"]["effective_facts"]
        first["test_input"] = ""
        binding["test_input_fact_ids"] = []
        assert validate_case_fact_bindings(
            test_cases=[first], raw_bindings=[binding], authoritative_facts=facts,
            expected_module_name=first["module"],
        )
        first["test_input"] = cases[0]["test_input"]
        assert first["test_input"], "非空输入断言需要真实的非空样本"
        try:
            validate_case_fact_bindings(
                test_cases=[first], raw_bindings=[binding], authoritative_facts=facts,
                expected_module_name=first["module"],
            )
        except ValueError as exc:
            assert "test_input必须绑定至少一个" in str(exc)
        else:
            raise AssertionError("非空测试输入缺少事实绑定时必须拒绝")

        legacy_input = deepcopy(run.input_payload)
        legacy_input.pop("enable_context_compression", None)
        legacy_input["compress"] = False
        normalized = _normalize_workflow_input("test_generation", legacy_input)
        assert normalized["enable_context_compression"] is False
        assert "compress" not in normalized
        assert legacy_input["compress"] is False
        db.refresh(run)
        assert run.run_context == original_context and run.output_payload == original_output
        print(json.dumps({
            "run_id": run.id, "persisted_cases": len(cases), "validated_batches": len(records),
            "missing_fact_rejected": rejected_fact, "validation_did_not_mutate_input": True,
            "empty_input_contract_checked": True, "export_input_values_match": True,
            "display_result_matches_persistence": True, "legacy_alias_normalized_at_boundary": True,
            "database_unchanged": True,
        }, ensure_ascii=False))


if __name__ == "__main__":
    main()
